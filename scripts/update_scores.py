from pathlib import Path
import requests
from openpyxl import load_workbook

REPO_ROOT = Path(__file__).resolve().parents[1]
WORKBOOK_PATH = REPO_ROOT / "data" / "2026 Masters Draft & Scoreboard.xlsx"

TOURNAMENT_ID = "401811941"
ESPN_API_URL = f"https://site.api.espn.com/apis/site/v2/sports/golf/pga/scoreboard?tournamentId={TOURNAMENT_ID}"

HEADERS = {
    "User-Agent": "Mozilla/5.0"
}

def clean(value):
    if value is None:
        return ""
    return str(value).strip()

def normalize_score(comp):
    score = clean(comp.get("score"))
    status = comp.get("status", {}) or {}
    status_display = clean(status.get("displayValue")).upper()
    status_short = clean(status.get("shortDisplayName")).upper()
    thru = clean(comp.get("currentHole")).upper()
    pos = clean(comp.get("curatedRank", {}).get("displayValue")).upper()

    if "CUT" in {score.upper(), status_display, status_short, thru, pos}:
        return "CUT"
    if "WD" in {score.upper(), status_display, status_short, thru, pos}:
        return "WD"
    if "DQ" in {score.upper(), status_display, status_short, thru, pos}:
        return "DQ"

    return score

def normalize_thru(comp):
    thru = clean(comp.get("currentHole"))
    if thru:
        return thru

    status = comp.get("status", {}) or {}
    status_display = clean(status.get("displayValue")).upper()
    status_short = clean(status.get("shortDisplayName")).upper()

    if "CUT" in {status_display, status_short}:
        return "CUT"
    if "WD" in {status_display, status_short}:
        return "WD"
    if "DQ" in {status_display, status_short}:
        return "DQ"

    return clean(status.get("displayValue") or status.get("shortDisplayName"))

def parse_competitor(comp):
    athlete = comp.get("athlete", {}) or {}
    linescores = comp.get("linescores", []) or []
    statistics = comp.get("statistics", []) or []

    player = clean(athlete.get("displayName"))
    pos = clean(comp.get("curatedRank", {}).get("displayValue") or comp.get("order"))
    score = normalize_score(comp)
    today = clean(comp.get("toPar"))
    thru = normalize_thru(comp)

    r1 = clean(linescores[0].get("value")) if len(linescores) > 0 else ""
    r2 = clean(linescores[1].get("value")) if len(linescores) > 1 else ""
    r3 = clean(linescores[2].get("value")) if len(linescores) > 2 else ""
    r4 = clean(linescores[3].get("value")) if len(linescores) > 3 else ""

    if pos.isdigit():
        tied = False
        for stat in statistics:
            name = clean(stat.get("name")).lower()
            display = clean(stat.get("displayValue")).lower()
            if name == "tied" and display == "true":
                tied = True
                break
        if tied:
            pos = f"T{pos}"

    return {
        "pos": pos,
        "player": player,
        "score": score,
        "today": today,
        "thru": thru,
        "r1": r1,
        "r2": r2,
        "r3": r3,
        "r4": r4,
    }

def fetch_scores():
    response = requests.get(ESPN_API_URL, headers=HEADERS, timeout=30)
    response.raise_for_status()
    data = response.json()

    rows = []

    events = data.get("events", []) or []
    for event in events:
        competitions = event.get("competitions", []) or []
        for competition in competitions:
            competitors = competition.get("competitors", []) or []
            for comp in competitors:
                row = parse_competitor(comp)
                if row["player"]:
                    rows.append(row)

    if not rows:
        raise RuntimeError("No golfers returned from ESPN API.")

    def sort_key(row):
        pos = row["pos"]
        if pos.startswith("T") and pos[1:].isdigit():
            return (0, int(pos[1:]), row["player"])
        if pos.isdigit():
            return (0, int(pos), row["player"])
        if pos in {"CUT", "WD", "DQ"}:
            return (2, 9999, row["player"])
        return (1, 9998, row["player"])

    rows.sort(key=sort_key)
    return rows

def update_excel(scores):
    wb = load_workbook(WORKBOOK_PATH)
    ws = wb["Scores"]

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=10):
        for cell in row:
            cell.value = None

    for row_index, item in enumerate(scores, start=2):
        ws.cell(row=row_index, column=1, value="")
        ws.cell(row=row_index, column=2, value=item["pos"])
        ws.cell(row=row_index, column=3, value=item["player"])
        ws.cell(row=row_index, column=4, value=item["score"])
        ws.cell(row=row_index, column=5, value=item["today"])
        ws.cell(row=row_index, column=6, value=item["thru"])
        ws.cell(row=row_index, column=7, value=item["r1"])
        ws.cell(row=row_index, column=8, value=item["r2"])
        ws.cell(row=row_index, column=9, value=item["r3"])
        ws.cell(row=row_index, column=10, value=item["r4"])

    wb.save(WORKBOOK_PATH)

def main():
    scores = fetch_scores()
    update_excel(scores)
    print(f"Updated workbook with {len(scores)} golfers.")

if __name__ == "__main__":
    main()
